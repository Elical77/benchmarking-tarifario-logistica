import os
import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


ARCHIVO_CDS = "contado.xlsx"
ARCHIVO_VC = "via_cargo_resultados_ok.csv"
ARCHIVO_OCA = "oca_resultados_ok.csv"
ARCHIVO_ANDREANI = "andreani_resultados_ok.csv"

SALIDA_EXCEL = "comparativa_precios.xlsx"
HISTORICO_CSV = "historico_powerbi.csv"
HISTORICO_XLSX = "historico_powerbi.xlsx"

HOJA_CDS = "cuadro01"

# IMPORTANTE: este relevamiento corresponde a mayo 2026
MES_RELEVAMIENTO = "2026-05"
MES_INICIO_HISTORICO = "2026-05"
ORIGEN = "Buenos Aires"

IVA = 1.21


COLORES = {
    "base": "1F4E78",
    "via_cargo": "548235",
    "andreani": "C00000",
    "oca": "7030A0",
    "control": "666666",
}


MAPEO_DESTINOS_CDS = {
    "bahia blanca": "bahia blanca",
    "bariloche": "bariloche",
    "comodoro rivadavia": "com. rivadavia",
    "cordoba": "cordoba",
    "corrientes": "corrientes",
    "mar del plata": "mar del plata",
    "neuquen": "neuquen",
    "resistencia": "resistencia",
    "rio gallegos": "rio gallegos",
    "salta": "salta",
    "tucuman": "tucuman",
}


def normalizar_texto(valor):
    texto = str(valor).strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def leer_cds():
    """
    Lee contado.xlsx, hoja cuadro01.

    La tabla de Cruz del Sur está por rangos/tope de KG.
    Ejemplo:
    - envío de 12 kg usa tarifa tope 15 kg
    - envío de 7 kg usa tarifa tope 10 kg
    """
    df = pd.read_excel(ARCHIVO_CDS, sheet_name=HOJA_CDS, header=None)

    fila_header = 8  # Fila 9 de Excel
    col_tope = 0     # Columna A
    col_unidad = 1   # Columna B

    destinos = {}

    for col in range(2, df.shape[1]):
        destino = df.iloc[fila_header, col]

        if pd.isna(destino):
            continue

        destino_key = normalizar_texto(destino)

        if destino_key:
            destinos[destino_key] = col

    registros = []

    for row in range(fila_header + 1, df.shape[0]):
        tope_kg = df.iloc[row, col_tope]
        unidad = df.iloc[row, col_unidad]

        if pd.isna(tope_kg) or pd.isna(unidad):
            continue

        if normalizar_texto(unidad) != "kg":
            continue

        try:
            tope_kg_num = float(tope_kg)
        except Exception:
            continue

        for destino_key, col in destinos.items():
            precio = df.iloc[row, col]

            if pd.isna(precio):
                continue

            try:
                precio_num = float(precio)
            except Exception:
                continue

            registros.append({
                "destino_cds_key": destino_key,
                "tope_cds_kg": tope_kg_num,
                "precio_cruz_del_sur": precio_num,
            })

    cds = pd.DataFrame(registros)

    if cds.empty:
        return cds

    # Si por algún motivo hay duplicado por destino/tope, tomamos el menor valor.
    cds = cds.groupby(
        ["destino_cds_key", "tope_cds_kg"],
        as_index=False
    )["precio_cruz_del_sur"].min()

    return cds


def buscar_precio_cds_por_rango(cds, destino_cds_key, kg):
    """
    Busca la tarifa de Cruz del Sur usando el menor tope de KG >= kg cotizado.
    """
    if pd.isna(destino_cds_key):
        return pd.Series({
            "precio_cruz_del_sur": None,
            "tope_cds_kg": None
        })

    candidatos = cds[
        (cds["destino_cds_key"] == destino_cds_key) &
        (cds["tope_cds_kg"] >= float(kg))
    ].copy()

    if candidatos.empty:
        return pd.Series({
            "precio_cruz_del_sur": None,
            "tope_cds_kg": None
        })

    candidato = candidatos.sort_values("tope_cds_kg").iloc[0]

    return pd.Series({
        "precio_cruz_del_sur": candidato["precio_cruz_del_sur"],
        "tope_cds_kg": candidato["tope_cds_kg"]
    })


def leer_via_cargo():
    df = pd.read_csv(ARCHIVO_VC, sep=";", encoding="utf-8-sig")
    df["destino_key"] = df["destino"].apply(normalizar_texto)
    df["kg"] = df["kg"].astype(float)

    df["precio_via_cargo"] = pd.to_numeric(
        df["via_cargo_agencia_agencia"],
        errors="coerce"
    )

    df = df[[
        "destino",
        "destino_key",
        "kg",
        "precio_via_cargo",
        "estado",
    ]].rename(columns={"estado": "estado_via_cargo"})

    df = df.drop_duplicates(subset=["destino_key", "kg"])

    return df


def leer_andreani():
    df = pd.read_csv(ARCHIVO_ANDREANI, sep=";", encoding="utf-8-sig")
    df["destino_key"] = df["destino"].apply(normalizar_texto)
    df["kg"] = df["kg"].astype(float)

    df["precio_andreani_bruto"] = pd.to_numeric(
        df["precio_andreani_sucursal"],
        errors="coerce"
    )

    # Andreani informa precios con IVA incluido.
    # Para comparar contra Cruz del Sur, neteamos IVA.
    df["precio_andreani"] = df["precio_andreani_bruto"] / IVA

    df = df[[
        "destino",
        "destino_key",
        "kg",
        "precio_andreani",
        "precio_andreani_bruto",
        "estado",
    ]].rename(columns={"estado": "estado_andreani"})

    df = df.drop_duplicates(subset=["destino_key", "kg"])

    return df


def leer_oca():
    df = pd.read_csv(ARCHIVO_OCA, sep=";", encoding="utf-8-sig")
    df["destino_key"] = df["destino"].apply(normalizar_texto)
    df["kg"] = df["kg"].astype(float)

    df["precio_oca_bruto"] = pd.to_numeric(
        df["precio_oca_sucursal"],
        errors="coerce"
    )

    # OCA publica precios con IVA incluido.
    # Para comparar contra Cruz del Sur, neteamos IVA.
    df["precio_oca"] = df["precio_oca_bruto"] / IVA

    df = df[[
        "destino",
        "destino_key",
        "kg",
        "zona_oca",
        "precio_oca",
        "precio_oca_bruto",
        "estado",
    ]].rename(columns={"estado": "estado_oca"})

    df = df.drop_duplicates(subset=["destino_key", "kg"])

    return df


def generar_tabla_comparativa():
    cds = leer_cds()
    vc = leer_via_cargo()
    andreani = leer_andreani()
    oca = leer_oca()

    # Base de comparación: misma matriz de destinos/kg que usamos en Vía Cargo.
    comp = vc.copy()
    comp["destino_cds_key"] = comp["destino_key"].map(MAPEO_DESTINOS_CDS)

    # Cruz del Sur por rango
    cds_lookup = comp.apply(
        lambda row: buscar_precio_cds_por_rango(
            cds,
            row["destino_cds_key"],
            row["kg"]
        ),
        axis=1
    )

    comp = pd.concat([comp, cds_lookup], axis=1)

    # Andreani
    comp = comp.merge(
        andreani[[
            "destino_key",
            "kg",
            "precio_andreani",
            "precio_andreani_bruto",
            "estado_andreani",
        ]],
        on=["destino_key", "kg"],
        how="left"
    )

    # OCA
    comp = comp.merge(
        oca[[
            "destino_key",
            "kg",
            "zona_oca",
            "precio_oca",
            "precio_oca_bruto",
            "estado_oca",
        ]],
        on=["destino_key", "kg"],
        how="left"
    )

    # Desvíos contra Cruz del Sur
    comp["desvio_vc_$"] = comp["precio_via_cargo"] - comp["precio_cruz_del_sur"]
    comp["desvio_vc_%"] = comp["desvio_vc_$"] / comp["precio_cruz_del_sur"]

    comp["desvio_andreani_$"] = comp["precio_andreani"] - comp["precio_cruz_del_sur"]
    comp["desvio_andreani_%"] = comp["desvio_andreani_$"] / comp["precio_cruz_del_sur"]

    comp["desvio_oca_$"] = comp["precio_oca"] - comp["precio_cruz_del_sur"]
    comp["desvio_oca_%"] = comp["desvio_oca_$"] / comp["precio_cruz_del_sur"]

    comp["estado_cruce"] = "OK"

    comp.loc[comp["precio_cruz_del_sur"].isna(), "estado_cruce"] = "SIN PRECIO CDS"
    comp.loc[comp["precio_via_cargo"].isna(), "estado_cruce"] = "SIN PRECIO VIA CARGO"
    comp.loc[comp["precio_andreani"].isna(), "estado_cruce"] = "SIN PRECIO ANDREANI"
    comp.loc[comp["precio_oca"].isna(), "estado_cruce"] = "SIN PRECIO OCA"

    comp = comp[[
        "destino",
        "kg",
        "precio_cruz_del_sur",

        "precio_via_cargo",
        "desvio_vc_$",
        "desvio_vc_%",

        "precio_andreani",
        "desvio_andreani_$",
        "desvio_andreani_%",

        "precio_oca",
        "desvio_oca_$",
        "desvio_oca_%",

        "zona_oca",
        "tope_cds_kg",
        "precio_andreani_bruto",
        "precio_oca_bruto",
        "estado_via_cargo",
        "estado_andreani",
        "estado_oca",
        "estado_cruce",
    ]]

    comp = comp.sort_values(["destino", "kg"]).reset_index(drop=True)

    return comp


def generar_base_powerbi(comp):
    registros = []

    for _, row in comp.iterrows():
        destino = row["destino"]
        kg = row["kg"]

        empresas = [
            ("Cruz del Sur", row["precio_cruz_del_sur"], "Agencia-Agencia", "Tarifario contado"),
            ("Vía Cargo", row["precio_via_cargo"], "Agencia-Agencia", "Cotizador online"),
            ("Andreani", row["precio_andreani"], "Sucursal neto IVA", "Cotizador online"),
            ("OCA", row["precio_oca"], "Sucursal neto IVA", "PDF tarifario"),
        ]

        for empresa, precio, tipo_servicio, fuente in empresas:
            registros.append({
                "mes": MES_RELEVAMIENTO,
                "origen": ORIGEN,
                "destino": destino,
                "kg": kg,
                "empresa": empresa,
                "precio": precio,
                "tipo_servicio": tipo_servicio,
                "fuente": fuente,
            })

    base_powerbi = pd.DataFrame(registros)
    base_powerbi["precio"] = pd.to_numeric(base_powerbi["precio"], errors="coerce").round(2)

    return base_powerbi


def generar_resumen(comp):
    ok = comp[comp["estado_cruce"] == "OK"].copy()

    resumen = ok.groupby("destino").agg(
        casos=("kg", "count"),
        cds_promedio=("precio_cruz_del_sur", "mean"),
        via_cargo_promedio=("precio_via_cargo", "mean"),
        andreani_promedio=("precio_andreani", "mean"),
        oca_promedio=("precio_oca", "mean"),
        desvio_vc_promedio=("desvio_vc_%", "mean"),
        desvio_andreani_promedio=("desvio_andreani_%", "mean"),
        desvio_oca_promedio=("desvio_oca_%", "mean"),
    ).reset_index()

    return resumen


def actualizar_historico(base_powerbi):
    """
    Actualiza el histórico acumulado.
    Si ya existe el mes, lo reemplaza para no duplicar.
    El histórico arranca en MES_INICIO_HISTORICO.
    """

    if os.path.exists(HISTORICO_XLSX):
        hist = pd.read_excel(HISTORICO_XLSX, sheet_name="Historico")
        hist = hist[hist["mes"] != MES_RELEVAMIENTO]
        hist = pd.concat([hist, base_powerbi], ignore_index=True)

    elif os.path.exists(HISTORICO_CSV):
        hist = pd.read_csv(HISTORICO_CSV, sep=";", encoding="utf-8-sig", decimal=",")
        hist = hist[hist["mes"] != MES_RELEVAMIENTO]
        hist = pd.concat([hist, base_powerbi], ignore_index=True)

    else:
        hist = base_powerbi.copy()

    # El histórico oficial arranca en mayo 2026.
    # Esto evita que queden corridas accidentales anteriores, como 2026-04.
    hist = hist[hist["mes"].astype(str) >= MES_INICIO_HISTORICO].copy()

    # Quitamos filas sin precio porque no sirven para Power BI.
    hist["precio"] = pd.to_numeric(hist["precio"], errors="coerce").round(2)
    hist = hist[hist["precio"].notna()].copy()

    hist.to_excel(
        HISTORICO_XLSX,
        sheet_name="Historico",
        index=False
    )

    hist.to_csv(
        HISTORICO_CSV,
        sep=";",
        index=False,
        encoding="utf-8-sig",
        decimal=","
    )


def escribir_excel(comp, base_powerbi, resumen):
    with pd.ExcelWriter(SALIDA_EXCEL, engine="openpyxl") as writer:
        comp.to_excel(writer, sheet_name="Comparativa Mes", index=False, startrow=2)
        base_powerbi.to_excel(writer, sheet_name="Base PowerBI", index=False)
        resumen.to_excel(writer, sheet_name="Resumen", index=False)

    wb = load_workbook(SALIDA_EXCEL)

    ws = wb["Comparativa Mes"]

    # Títulos agrupados
    ws.merge_cells("A1:B1")
    ws["A1"] = "Base"

    ws["C1"] = "Cruz del Sur"

    ws.merge_cells("D1:F1")
    ws["D1"] = "Vía Cargo"

    ws.merge_cells("G1:I1")
    ws["G1"] = "Andreani neto IVA"

    ws.merge_cells("J1:L1")
    ws["J1"] = "OCA neto IVA"

    ws.merge_cells("M1:T1")
    ws["M1"] = "Control"

    grupos = {
        "A1": COLORES["base"],
        "C1": COLORES["base"],
        "D1": COLORES["via_cargo"],
        "G1": COLORES["andreani"],
        "J1": COLORES["oca"],
        "M1": COLORES["control"],
    }

    for cell_ref, color in grupos.items():
        cell = ws[cell_ref]
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados fila 3
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[3]:
        cell.fill = PatternFill("solid", fgColor=COLORES["base"])
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ["D", "E", "F"]:
        ws[f"{col}3"].fill = PatternFill("solid", fgColor=COLORES["via_cargo"])

    for col in ["G", "H", "I"]:
        ws[f"{col}3"].fill = PatternFill("solid", fgColor=COLORES["andreani"])

    for col in ["J", "K", "L"]:
        ws[f"{col}3"].fill = PatternFill("solid", fgColor=COLORES["oca"])

    for col in ["M", "N", "O", "P", "Q", "R", "S", "T"]:
        ws[f"{col}3"].fill = PatternFill("solid", fgColor=COLORES["control"])

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Formatos
    money_cols = ["C", "D", "E", "G", "H", "J", "K", "O", "P"]
    pct_cols = ["F", "I", "L"]

    for col in money_cols:
        for cell in ws[col][3:]:
            cell.number_format = '$ #,##0'

    for col in pct_cols:
        for cell in ws[col][3:]:
            cell.number_format = '0%'

    for col in ["B", "N"]:
        for cell in ws[col][3:]:
            cell.number_format = '0'

    # Condicional: negativos rojo, positivos verde
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="006100")

    for col in ["F", "I", "L"]:
        rango = f"{col}4:{col}{ws.max_row}"
        ws.conditional_formatting.add(
            rango,
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
        )
        ws.conditional_formatting.add(
            rango,
            CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=green_font)
        )

    # Anchos
    widths = {
        "A": 22,
        "B": 8,
        "C": 16,
        "D": 16,
        "E": 14,
        "F": 12,
        "G": 16,
        "H": 14,
        "I": 12,
        "J": 16,
        "K": 14,
        "L": 12,
        "M": 14,
        "N": 12,
        "O": 18,
        "P": 16,
        "Q": 16,
        "R": 18,
        "S": 14,
        "T": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Formato hojas secundarias
    for sheet_name in ["Base PowerBI", "Resumen"]:
        ws2 = wb[sheet_name]
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions

        for cell in ws2[1]:
            cell.fill = PatternFill("solid", fgColor=COLORES["base"])
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_cells in enumerate(ws2.columns, start=1):
            max_len = 0
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 35)

        for row in ws2.iter_rows():
            for cell in row:
                cell.border = border

    wb.save(SALIDA_EXCEL)


def main():
    comp = generar_tabla_comparativa()
    base_powerbi = generar_base_powerbi(comp)
    resumen = generar_resumen(comp)

    escribir_excel(comp, base_powerbi, resumen)
    actualizar_historico(base_powerbi)

    duplicados = comp.duplicated(subset=["destino", "kg"]).sum()

    print("Proceso terminado.")
    print(f"Archivo Excel generado: {SALIDA_EXCEL}")
    print(f"Histórico Power BI actualizado: {HISTORICO_XLSX}")
    print(f"CSV auxiliar actualizado: {HISTORICO_CSV}")
    print(f"Filas comparativa: {len(comp)}")
    print(f"Duplicados destino + kg: {duplicados}")
    print("Estados del cruce:")
    print(comp["estado_cruce"].value_counts())


if __name__ == "__main__":
    main()