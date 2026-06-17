import csv
import math
import re
import unicodedata
from openpyxl import load_workbook


INPUT_EXCEL = "Comparativo Cotizacion - Febrero 2026.xlsm"
SHEET_NAME = "Analisis"

OUTPUT_INPUT = "via_cargo_input.csv"
OUTPUT_EXCLUIDOS = "via_cargo_excluidos.csv"

ORIGEN_CP_DEFAULT = "1303"
VALOR_DECLARADO_DEFAULT = 75000

# Primera pasada segura:
# Vía Cargo informa límite por caja/paquete. Para evitar errores, excluimos >50 kg por ahora.
EXCLUIR_MAYORES_A_50KG = True
EXCLUIR_AFORO = True


CP_DESTINOS = {
    "bahia blanca": "8000",
    "bariloche": "8400",
    "comodoro rivadavia": "9000",
    "neuquen": "8300",
    "rio gallegos": "9400",
    "salta": "4400",
    "tucuman": "4000",
    "mar del plata": "7600",
    "resistencia": "3500",
    "corrientes": "3400",
    "cordoba": "5000",
    "rosario": "2000",
    "mendoza": "5500",
}


def normalizar_texto(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def limpiar_destino(destino):
    """
    Quita la palabra Aforo para mapear CP.
    Ejemplo: 'Bahía Blanca Aforo' -> 'Bahía Blanca'
    """
    texto = str(destino).strip()
    texto = re.sub(r"\s+aforo\s*$", "", texto, flags=re.IGNORECASE)
    return texto.strip()


def leer_filas_excel():
    wb = load_workbook(INPUT_EXCEL, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]

    filas = []
    origen_actual = None

    for row in ws.iter_rows(min_row=7, values_only=True):
        origen = row[0]
        destino = row[1]
        kg = row[2]

        if origen not in [None, ""]:
            origen_actual = origen

        if destino in [None, ""] or kg in [None, ""]:
            continue

        destino_txt = str(destino).strip()

        if normalizar_texto(destino_txt) in ["destino", "promedio"]:
            continue

        if normalizar_texto(origen_actual) in ["origen", "base", ""]:
            origen_actual = "Buenos Aires"

        try:
            kg_num = float(kg)
        except Exception:
            continue

        filas.append({
            "origen": str(origen_actual).strip() if origen_actual else "Buenos Aires",
            "destino": destino_txt,
            "kg": kg_num,
        })

    return filas


def generar_input():
    filas_excel = leer_filas_excel()

    filas_ok = []
    filas_excluidas = []
    vistos = set()

    for fila in filas_excel:
        destino_original = fila["destino"]
        destino_limpio = limpiar_destino(destino_original)
        destino_key = normalizar_texto(destino_limpio)
        kg = fila["kg"]

        motivo_exclusion = ""

        if EXCLUIR_AFORO and "aforo" in normalizar_texto(destino_original):
            motivo_exclusion = "Fila Aforo excluida en primera pasada"

        elif EXCLUIR_MAYORES_A_50KG and kg > 50:
            motivo_exclusion = "Kg mayor a 50 excluido en primera pasada"

        elif destino_key not in CP_DESTINOS:
            motivo_exclusion = f"No hay CP mapeado para destino: {destino_original}"

        clave = (destino_key, kg)

        if motivo_exclusion:
            filas_excluidas.append({
                "origen": fila["origen"],
                "destino": destino_original,
                "kg": kg,
                "motivo": motivo_exclusion,
            })
            continue

        if clave in vistos:
            continue

        vistos.add(clave)

        filas_ok.append({
            "origen_cp": ORIGEN_CP_DEFAULT,
            "destino_cp": CP_DESTINOS[destino_key],
            "destino": destino_limpio,
            "bultos": 1,
            "kg": int(kg) if kg.is_integer() else kg,
            "valor_declarado": VALOR_DECLARADO_DEFAULT,
        })

    filas_ok.sort(key=lambda x: (x["destino"], float(x["kg"])))

    with open(OUTPUT_INPUT, "w", newline="", encoding="utf-8-sig") as archivo:
        columnas = ["origen_cp", "destino_cp", "destino", "bultos", "kg", "valor_declarado"]
        writer = csv.DictWriter(archivo, fieldnames=columnas)
        writer.writeheader()
        writer.writerows(filas_ok)

    with open(OUTPUT_EXCLUIDOS, "w", newline="", encoding="utf-8-sig") as archivo:
        columnas = ["origen", "destino", "kg", "motivo"]
        writer = csv.DictWriter(archivo, fieldnames=columnas)
        writer.writeheader()
        writer.writerows(filas_excluidas)

    print("Proceso terminado.")
    print(f"Filas generadas para Vía Cargo: {len(filas_ok)}")
    print(f"Filas excluidas/revisar: {len(filas_excluidas)}")
    print(f"Archivo generado: {OUTPUT_INPUT}")
    print(f"Archivo de excluidos: {OUTPUT_EXCLUIDOS}")


if __name__ == "__main__":
    generar_input()